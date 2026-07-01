"""
Attendance Engine — Parts 1 & 2 (Seating Plan Ingestion Edition)
=================================================================

Loads, validates, and normalizes:
  - Part 1: Master class roster from a multi-sheet rollcall workbook.
  - Part 2: Wide-format seating plan attendance file into long-format batches.

The cleaned DataFrames produced here serve as the foundation for de-duplication,
logging (Part 3), aggregations (Part 4), and frontend rendering (Part 5).
"""

from __future__ import annotations

import io
import os
from datetime import date, datetime
from pathlib import Path

import pandas as pd

from config import (
    DEFAULT_COLUMN_MAPPING,
    SEATING_ENROLL_HEADER,
    SEATING_BOUNDARY_HEADER,
    SLOT_MAP,
    STATUS_MAP,
    DATA_DIR,
    MASTER_LIST_FILENAME,
    SEATING_PLAN_FILENAME,
)


# ═══════════════════════════════════════════════════════════════════════════
#  Part 1 — Rollcall Master Roster Loader
# ═══════════════════════════════════════════════════════════════════════════

def load_master_roster(
    filepath: str | Path,
    column_mapping: dict[str, str] | None = None,
) -> pd.DataFrame:
    """Load and clean the master roster from a multi-sheet Excel workbook.

    Each sheet represents a division, where the sheet name is the division.

    Parameters
    ----------
    filepath : str or Path
        Path to the multi-sheet Excel workbook containing the roster.
    column_mapping : dict, optional
        Maps internal fields ('enrollment_number', 'name', 'roll_number',
        'department', 'semester') to actual column headers in the sheets.

    Returns
    -------
    pd.DataFrame
        Cleaned master roster indexed by enrollment_number, with columns:
        name, division, roll_number, department, semester.

    Raises
    ------
    FileNotFoundError
        If the file doesn't exist.
    KeyError
        If any sheet lacks a required mapped column.
    ValueError
        If there are empty sheets, duplicates, or nulls in enrollment_number.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Master roster workbook not found: {filepath}")

    mapping = column_mapping or DEFAULT_COLUMN_MAPPING
    required_keys = ["enrollment_number", "name", "roll_number", "department", "semester"]

    # Verify column mapping dictionary keys
    for rk in required_keys:
        if rk not in mapping:
            raise KeyError(f"Missing required key in column mapping: '{rk}'")

    excel_file = pd.ExcelFile(filepath, engine="openpyxl")
    sheet_names = excel_file.sheet_names

    dfs = []
    for sheet in sheet_names:
        df_sheet = excel_file.parse(sheet)
        
        # Verify columns exist in sheet
        for key in required_keys:
            excel_col = mapping[key]
            if excel_col not in df_sheet.columns:
                raise KeyError(
                    f"Column '{excel_col}' not found in sheet '{sheet}'. "
                    f"Available columns: {list(df_sheet.columns)}"
                )

        # Rename columns to internal names and extract only required columns
        inverse_map = {mapping[k]: k for k in required_keys}
        df_sheet = df_sheet.rename(columns=inverse_map)
        df_sheet = df_sheet[required_keys].copy()
        
        # Record the sheet name as division
        df_sheet["division"] = sheet
        
        dfs.append(df_sheet)

    if not dfs:
        raise ValueError(f"Roster workbook '{filepath}' contains no sheets.")

    df = pd.concat(dfs, ignore_index=True)

    # ------------------------------------------------------------------
    # Roster-wide uniqueness and null checks
    # ------------------------------------------------------------------
    # Null Check
    null_mask = df["enrollment_number"].isna()
    if null_mask.any():
        raise ValueError(
            f"Enrollment number column contains {null_mask.sum()} null value(s). "
            f"Every student record must have a valid enrollment number."
        )

    # Duplicate Check
    dup_mask = df["enrollment_number"].astype(str).str.strip().duplicated(keep=False)
    if dup_mask.any():
        dups = sorted(df.loc[dup_mask, "enrollment_number"].unique())
        raise ValueError(
            f"Duplicate enrollment numbers found in roster workbook: {dups}. "
            f"Enrollment number must be unique across all sheets."
        )

    # ------------------------------------------------------------------
    # Normalization
    # ------------------------------------------------------------------
    df["enrollment_number"] = df["enrollment_number"].astype(str).str.strip()
    df["name"] = df["name"].astype(str).str.strip()
    df["roll_number"] = df["roll_number"].astype(str).str.strip()
    df["department"] = df["department"].astype(str).str.strip().str.upper()
    df["division"] = df["division"].astype(str).str.strip().str.upper()
    df["semester"] = df["semester"].astype(str).str.strip().str.upper()

    return df.set_index("enrollment_number")


# ═══════════════════════════════════════════════════════════════════════════
#  Part 2, Path B — Manual Wide File Ingestion
# ═══════════════════════════════════════════════════════════════════════════

def load_attendance_from_file(
    filepath: str | Path,
    master_roster: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """Load seating plan attendance from a local Excel path and normalize it.

    Parameters
    ----------
    filepath : str or Path
        Path to the seating plan Excel file.
    master_roster : pd.DataFrame
        The master roster DataFrame from load_master_roster.

    Returns
    -------
    dict
        Contains DataFrames: 'attendance' and 'unmatched'.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Seating plan file not found: {path}")
    raw_bytes = path.read_bytes()
    return normalize_attendance(raw_bytes, master_roster, filename=path.name)


# ═══════════════════════════════════════════════════════════════════════════
#  Part 2, Shared Parsing & Wide-To-Long Unpivoting
# ═══════════════════════════════════════════════════════════════════════════

def normalize_attendance(
    raw_data: bytes,
    master_roster: pd.DataFrame,
    filename: str = "attendance.xlsx",
) -> dict[str, pd.DataFrame]:
    """Parse wide training seating plan Excel data into a clean, long-format DataFrame.

    Converts repeating date columns into vertical database-ready rows.
    Handles merged header columns correctly by forward-filling dates from row 1,
    and pairing them with slots from row 2.

    Parameters
    ----------
    raw_data : bytes
        Excel file content bytes.
    master_roster : pd.DataFrame
        The master roster DataFrame (Part 1, indexed by enrollment_number).
    filename : str
        Name of the file (used for choosing openpyxl parser).

    Returns
    -------
    dict
        "attendance" : pd.DataFrame
            Columns: enrollment_number, session_date, session_slot, status.
            Only includes students matched with enrollment_number in master_roster.
        "unmatched" : pd.DataFrame
            Columns: enrollment_number, session_date, session_slot, status.
            Holds entries whose enrollment numbers do not appear in master_roster.

    Raises
    ------
    ValueError
        If headers are missing, dates/slots cannot be parsed, or if any status
        value cannot be mapped to Present, Absent, or Leave.
    KeyError
        If required identify columns are missing from the sheet.
    """
    buf = io.BytesIO(raw_data)
    
    # Read sheet without header parser to inspect rows directly
    raw_df = pd.read_excel(buf, header=None, engine="openpyxl")
    
    if raw_df.shape[0] < 2:
        raise ValueError("Seating plan Excel sheet must have at least two header rows.")

    hdr1 = raw_df.iloc[0].tolist()
    hdr2 = raw_df.iloc[1].tolist()

    # 1. Identify enrollment column and boundary
    enroll_col_idx = None
    boundary_col_idx = None

    for idx, val in enumerate(hdr1):
        if not isinstance(val, str):
            continue
        cleaned = val.strip().lower()
        if cleaned == SEATING_ENROLL_HEADER.lower():
            enroll_col_idx = idx
        if cleaned == SEATING_BOUNDARY_HEADER.lower():
            boundary_col_idx = idx

    if enroll_col_idx is None:
        raise KeyError(f"Required column '{SEATING_ENROLL_HEADER}' not found in header row 1.")
    if boundary_col_idx is None:
        raise KeyError(f"Required column '{SEATING_BOUNDARY_HEADER}' not found in header row 1.")

    # Fixed info ends at boundary, repeating date blocks start right after
    start_date_col = boundary_col_idx + 1

    if start_date_col >= raw_df.shape[1]:
        raise ValueError("No date-block attendance columns detected after fixed boundary.")

    # 2. Extract and forward-fill date labels
    raw_dates = hdr1[start_date_col:]
    dates_filled = pd.Series(raw_dates).ffill().tolist()

    # 3. Compile date columns configuration
    # Choice: Slot 1 maps to Morning, Slot 2 maps to Evening
    date_blocks = []
    for offset, raw_date in enumerate(dates_filled):
        col_idx = start_date_col + offset
        slot_raw = str(hdr2[col_idx]).strip()
        slot_key = slot_raw.lower()

        if slot_key not in SLOT_MAP:
            raise ValueError(
                f"Row 2 Col {col_idx + 1}: Unexpected slot value '{slot_raw}'. "
                f"Expected Slot 1/Slot 2 or Morning/Evening."
            )
        slot_name = SLOT_MAP[slot_key]

        # Standardize date to YYYY-MM-DD
        try:
            if hasattr(raw_date, "strftime"):
                date_str = raw_date.strftime("%Y-%m-%d")
            else:
                date_str = pd.to_datetime(str(raw_date).strip()).strftime("%Y-%m-%d")
        except Exception:
            raise ValueError(f"Row 1 Col {col_idx + 1}: Invalid date value '{raw_date}'.")

        date_blocks.append((col_idx, date_str, slot_name))

    # 4. Iterate and unpivot rows
    matched_records = []
    unmatched_records = []

    roster_enroll_set = set(master_roster.index)

    for r_idx in range(2, raw_df.shape[0]):
        enroll_raw = raw_df.iloc[r_idx, enroll_col_idx]
        if pd.isna(enroll_raw):
            continue

        enroll_no = str(enroll_raw).strip()
        if not enroll_no:
            continue

        is_matched = enroll_no in roster_enroll_set

        for col_idx, s_date, s_slot in date_blocks:
            status_raw = raw_df.iloc[r_idx, col_idx]
            
            # Treat missing values/NaN as invalid instead of silent pass
            if pd.isna(status_raw):
                status_raw = "MISSING"

            status_clean = str(status_raw).strip().lower()

            if status_clean not in STATUS_MAP:
                raise ValueError(
                    f"Row {r_idx + 1}: Student '{enroll_no}' has unmappable status '{status_raw}' "
                    f"for session on {s_date} ({s_slot})."
                )

            status = STATUS_MAP[status_clean]

            record = {
                "enrollment_number": enroll_no,
                "session_date":      s_date,
                "session_slot":      s_slot,
                "status":            status,
            }

            if is_matched:
                matched_records.append(record)
            else:
                unmatched_records.append(record)

    cols = ["enrollment_number", "session_date", "session_slot", "status"]
    attendance_df = pd.DataFrame(matched_records, columns=cols)
    unmatched_df = pd.DataFrame(unmatched_records, columns=cols)

    return {
        "attendance": attendance_df,
        "unmatched":  unmatched_df,
    }
# ═══════════════════════════════════════════════════════════════════════════
#  Part 3 — Append-Only Session Log
# ═══════════════════════════════════════════════════════════════════════════

# Columns in the persistent log CSV
_LOG_COLUMNS = [
    "enrollment_number",
    "status",
    "session_date",
    "session_slot",
    "logged_at",
]


def _default_log_path() -> Path:
    """Return default path to attendance_log.csv."""
    from config import ATTENDANCE_LOG_FILENAME
    return Path(DATA_DIR) / ATTENDANCE_LOG_FILENAME


def append_to_log(
    batch: pd.DataFrame,
    log_path: str | Path | None = None,
) -> dict[str, int]:
    """Append a normalized attendance batch to the persistent CSV log.

    De-duplication key: (enrollment_number, session_date, session_slot).

    - Exact duplicate (same key AND status) -> silently skipped.
    - Correction (same key, different status) -> updated in place and logged.
    - New row -> appended with a logged_at timestamp.
    """
    path = Path(log_path) if log_path else _default_log_path()
    os.makedirs(path.parent, exist_ok=True)

    if path.exists() and path.stat().st_size > 0:
        existing = pd.read_csv(path, dtype=str)
    else:
        existing = pd.DataFrame(columns=_LOG_COLUMNS)

    existing_keys = {}
    for idx, row in existing.iterrows():
        key = (row["enrollment_number"], row["session_date"], row["session_slot"])
        existing_keys[key] = idx

    now_str = datetime.now().isoformat(timespec="seconds")
    new_rows = []
    inserted = 0
    skipped = 0
    corrected = 0

    for _, row in batch.iterrows():
        key = (row["enrollment_number"], row["session_date"], row["session_slot"])

        if key in existing_keys:
            existing_idx = existing_keys[key]
            old_status = existing.at[existing_idx, "status"]
            new_status = row["status"]

            if old_status == new_status:
                skipped += 1
            else:
                print(
                    f"  [CORRECTION] {key[0]} | {key[1]} {key[2]}: "
                    f"{old_status} -> {new_status}"
                )
                existing.at[existing_idx, "status"] = new_status
                existing.at[existing_idx, "logged_at"] = now_str
                corrected += 1
        else:
            new_rows.append({
                "enrollment_number": row["enrollment_number"],
                "status":            row["status"],
                "session_date":      row["session_date"],
                "session_slot":      row["session_slot"],
                "logged_at":         now_str,
            })
            existing_keys[key] = -1
            inserted += 1

    if new_rows:
        new_df = pd.DataFrame(new_rows, columns=_LOG_COLUMNS)
        existing = pd.concat([existing, new_df], ignore_index=True)

    if inserted > 0 or corrected > 0:
        existing.to_csv(path, index=False)

    return {"inserted": inserted, "skipped": skipped, "corrected": corrected}


def load_attendance_log(
    log_path: str | Path | None = None,
) -> pd.DataFrame:
    """Load the full attendance log from CSV, parsing dates properly."""
    path = Path(log_path) if log_path else _default_log_path()

    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=_LOG_COLUMNS)

    df = pd.read_csv(path, dtype=str)
    # Parse session_date into datetime.date format
    df["session_date"] = pd.to_datetime(df["session_date"]).dt.date
    return df


def get_log_summary(
    log_path: str | Path | None = None,
) -> dict:
    """Return basic summary stats for the log."""
    df = load_attendance_log(log_path)

    if df.empty:
        return {
            "total_rows": 0,
            "total_sessions": 0,
            "earliest_date": None,
            "latest_date": None,
        }

    unique_sessions = df.groupby(["session_date", "session_slot"]).ngroups
    earliest = str(df["session_date"].min())
    latest = str(df["session_date"].max())

    return {
        "total_rows": len(df),
        "total_sessions": unique_sessions,
        "earliest_date": earliest,
        "latest_date": latest,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  Part 4 — Student Aggregation / Summary
# ═══════════════════════════════════════════════════════════════════════════

def compute_student_summary(
    master_roster: pd.DataFrame,
    log_df: pd.DataFrame,
) -> pd.DataFrame:
    """Produce a per-student summary matching all students in the master roster.

    Critical requirement: Left-joins master roster to log. Every row in the output
    retains roster properties (roll_number, division, department).
    Every student in the roster appears in the output, even those with zero
    sessions logged.

    Parameters
    ----------
    master_roster : pd.DataFrame
        Master roster DataFrame indexed by enrollment_number.
    log_df : pd.DataFrame
        Full attendance log DataFrame (columns include enrollment_number and status).

    Returns
    -------
    pd.DataFrame
        Columns: enrollment_number, name, department, division, roll_number,
                 total_sessions, present_count, absent_count, leave_count.
    """
    if log_df.empty:
        summary = pd.DataFrame(index=master_roster.index)
        summary["total_sessions"] = 0
        summary["present_count"] = 0
        summary["absent_count"] = 0
        summary["leave_count"] = 0
    else:
        # Ensure we only aggregate log entries belonging to active roster students
        valid_log = log_df[log_df["enrollment_number"].isin(master_roster.index)].copy()

        # Group count for total sessions
        total_sessions = valid_log.groupby("enrollment_number").size()

        # Pivot to get status counts per student
        status_counts = pd.crosstab(valid_log["enrollment_number"], valid_log["status"])
        
        # Ensure all standard columns are present
        for status in ["Present", "Absent", "Leave"]:
            if status not in status_counts.columns:
                status_counts[status] = 0

        # Construct the summary table matching the roster index structure
        summary = pd.DataFrame(index=master_roster.index)
        summary["total_sessions"] = total_sessions
        summary["present_count"] = status_counts["Present"]
        summary["absent_count"] = status_counts["Absent"]
        summary["leave_count"] = status_counts["Leave"]

    # Fill NaNs with 0 for students with no logs
    summary["total_sessions"] = summary["total_sessions"].fillna(0).astype(int)
    summary["present_count"] = summary["present_count"].fillna(0).astype(int)
    summary["absent_count"] = summary["absent_count"].fillna(0).astype(int)
    summary["leave_count"] = summary["leave_count"].fillna(0).astype(int)

    # Perform left join: Roster -> Summary
    joined = master_roster.join(summary, how="left")

    # Reset index to output enrollment_number as a column
    result = joined.reset_index()

    # Reorder columns as requested
    col_order = [
        "enrollment_number",
        "name",
        "department",
        "division",
        "roll_number",
        "total_sessions",
        "present_count",
        "absent_count",
        "leave_count",
    ]
    return result[col_order]


def generate_frontend_json(
    roster_path: str | Path | None = None,
    log_path: str | Path | None = None,
    output_path: str | Path | None = None,
) -> None:
    """Compile roster and log data into frontend/output/attendance_data.json."""
    import json

    r_path = Path(roster_path) if roster_path else Path(DATA_DIR) / MASTER_LIST_FILENAME
    l_path = Path(log_path) if log_path else _default_log_path()
    out_path = Path(output_path) if output_path else Path("frontend") / "output" / "attendance_data.json"

    # 1. Load roster & log
    roster_df = load_master_roster(r_path)
    log_df = load_attendance_log(l_path)

    # 2. Get unique sessions
    sessions_list = []
    if not log_df.empty:
        # Sort sessions chronologically
        sessions_meta_df = log_df[["session_date", "session_slot"]].drop_duplicates().sort_values(["session_date", "session_slot"])
        for _, row in sessions_meta_df.iterrows():
            d_str = str(row["session_date"])
            sessions_list.append({
                "date": d_str,
                "slot": row["session_slot"],
                "title": f"{row['session_slot']} Session"
            })

    # 3. Build helper lookup for fast queries
    log_lookup = {}
    if not log_df.empty:
        for _, row in log_df.iterrows():
            key = (row["enrollment_number"], str(row["session_date"]), row["session_slot"])
            log_lookup[key] = row["status"]

    # 4. Process students
    students_list = []
    for enroll_no, row in roster_df.iterrows():
        student_sessions = []
        present_count = 0
        total_sessions = len(sessions_list)

        for sess in sessions_list:
            s_date = sess["date"]
            s_slot = sess["slot"]
            s_title = sess["title"]

            # Look up status, default to "Absent" if not found in log
            status = log_lookup.get((enroll_no, s_date, s_slot), "Absent")
            student_sessions.append({
                "date": s_date,
                "title": s_title,
                "status": status.lower()  # lowercase present/absent/leave for frontend
            })

            if status.lower() == "present":
                present_count += 1

        pct = (present_count / total_sessions * 100) if total_sessions > 0 else 0.0

        students_list.append({
            "enrollment_no": enroll_no,
            "name":          row["name"],
            "department":    row["department"],
            "division":      row["division"],
            "roll_number":   row["roll_number"],
            "sessions":      student_sessions,
            "present_count": present_count,
            "total_sessions": total_sessions,
            "attendance_pct": pct
        })

    # 5. Build payload and write
    payload = {
        "generated_at": datetime.now().isoformat(),
        "sessions_meta": [
            {"date": s["date"], "title": s["title"]}
            for s in sessions_list
        ],
        "students": students_list
    }

    os.makedirs(out_path.parent, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"  [OK] Exported attendance_data.json to {out_path} ({len(students_list)} students, {len(sessions_list)} sessions)")


# ═══════════════════════════════════════════════════════════════════════════
#  Demo / self-test
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from openpyxl import Workbook
    
    SAMPLE_ROSTER = os.path.join(DATA_DIR, MASTER_LIST_FILENAME)
    SAMPLE_SEATING = os.path.join(DATA_DIR, SEATING_PLAN_FILENAME)
    LOG_PATH = os.path.join(DATA_DIR, "attendance_log_test.csv")

    print("=" * 70)
    print("  TEST 1 — Load Multi-sheet Master Roster (Happy Path)")
    print("=" * 70)
    try:
        roster = load_master_roster(SAMPLE_ROSTER)
        print(f"\n[OK] Loaded {len(roster)} students from master roster.\n")
        print(roster.to_string())
        print()
    except Exception as exc:
        print(f"\n[FAIL] Master roster loading failed: {exc}\n")

    # Happy path seating plan parse
    print("=" * 70)
    print("  TEST 2 — Ingest Seating Plan & Wide-to-Long Parse (Happy Path)")
    print("=" * 70)
    try:
        result = load_attendance_from_file(SAMPLE_SEATING, roster)
        att = result["attendance"]
        unm = result["unmatched"]
        print(f"\n[OK] Matched Attendance Records ({len(att)} rows):")
        print(att.to_string(index=False))
        print(f"\n[OK] Unmatched Attendance Records ({len(unm)} rows):")
        print(unm.to_string(index=False))
        print()
    except Exception as exc:
        print(f"\n[FAIL] Seating plan ingestion failed: {exc}\n")

    # Test 3: Unmatched Enrollment Number Check
    print("=" * 70)
    print("  TEST 3 — Unmatched Enrollment Numbers Detection")
    print("=" * 70)
    test_unmatched_path = os.path.join(DATA_DIR, "_test_unmatched.xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        fixed = ["Sr No", "Final Batch", "Floor", "Class", "Venue", "Enrollment No",
                 "Name", "Gender", "Course", "Branch", "College", "Year of Passing", "Parul Mail ID"]
        ws.append(fixed + ["2026-07-01"])
        ws.merge_cells("N1:O1")
        ws.append([""] * 13 + ["Slot 1", "Slot 2"])
        ws.append(["1", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-999", "Ghost Student", "Male", "B.Tech", "CS", "PIT", "2027", "ghost@parul.in",
                  "Present", "Absent"])
        wb.save(test_unmatched_path)

        res = load_attendance_from_file(test_unmatched_path, roster)
        unmatched_df = res["unmatched"]
        print(f"\n[OK] Successfully caught unmatched students:")
        print(unmatched_df.to_string(index=False))
        print()
    except Exception as exc:
        print(f"\n[FAIL] Unexpected error during unmatched test: {exc}\n")
    finally:
        if os.path.exists(test_unmatched_path):
            os.remove(test_unmatched_path)

    # Test 4: Malformed Status Check
    print("=" * 70)
    print("  TEST 4 — Malformed Status Mapping Check (Expect Error)")
    print("=" * 70)
    test_bad_status_path = os.path.join(DATA_DIR, "_test_bad_status.xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        fixed = ["Sr No", "Final Batch", "Floor", "Class", "Venue", "Enrollment No",
                 "Name", "Gender", "Course", "Branch", "College", "Year of Passing", "Parul Mail ID"]
        ws.append(fixed + ["2026-07-01"])
        ws.merge_cells("N1:O1")
        ws.append([""] * 13 + ["Slot 1", "Slot 2"])
        ws.append(["1", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-001", "Aarav Patel", "Male", "B.Tech", "CS", "PIT", "2027", "aarav@parul.in",
                  "XYZ", "Present"])
        wb.save(test_bad_status_path)

        load_attendance_from_file(test_bad_status_path, roster)
        print("\n[FAIL] Expected ValueError for malformed status, but none was raised.\n")
    except ValueError as exc:
        print(f"\n[OK] Correctly caught expected status error:\n  {exc}\n")
    except Exception as exc:
        print(f"\n[FAIL] Unexpected exception ({type(exc).__name__}): {exc}\n")
    finally:
        if os.path.exists(test_bad_status_path):
            os.remove(test_bad_status_path)

    # Test 5: SharePoint URL Conversion Demo
    print("=" * 70)
    print("  TEST 5 — SharePoint Link Fetch Conversion Demo (No Network)")
    print("=" * 70)
    from sharepoint_fetch import convert_share_link_to_download_url, SharePointURLError
    
    placeholder_link = "https://mytenant.sharepoint.com/:x:/s/ClassRoster/ETestLink12345"
    try:
        dl_url = convert_share_link_to_download_url(placeholder_link)
        print(f"\n[OK] Converted SharePoint link successfully:")
        print(f"  Source URL: {placeholder_link}")
        print(f"  Direct Download URL: {dl_url}\n")
    except SharePointURLError as exc:
        print(f"\n[FAIL] Link conversion failed: {exc}\n")

    # Test 6: Append to Log and Compute Aggregations
    print("=" * 70)
    print("  TEST 6 — Database Logging & Per-Student Summary (Part 4)")
    print("=" * 70)
    if os.path.exists(LOG_PATH):
        os.remove(LOG_PATH)
        
    try:
        # Append seating plan happy path batch to log
        append_to_log(att, log_path=LOG_PATH)
        
        # Load log back
        log_df = load_attendance_log(log_path=LOG_PATH)
        
        # Compute summary
        summary_df = compute_student_summary(roster, log_df)
        
        print("\n[OK] Per-Student Summary Stats:")
        print(summary_df.to_string(index=False))
        print()
        
        # Verify that ENR-004 and ENR-007 are present with 0 sessions
        zero_history_students = summary_df[summary_df["total_sessions"] == 0]["enrollment_number"].tolist()
        print(f"[OK] Students with zero attendance history: {zero_history_students}")
        if set(zero_history_students) == {"ENR-004", "ENR-007"}:
            print("  (Correct -- ENR-004 and ENR-007 correctly kept as 0 in summary)\n")
        else:
            print("  [FAIL] Incorrect listing of zero history students.\n")
            
    except Exception as exc:
        print(f"\n[FAIL] Summary aggregation failed: {exc}\n")
    finally:
        if os.path.exists(LOG_PATH):
            os.remove(LOG_PATH)

    print("=" * 70)
    print("  All Part 1 + Part 2 + Part 3 + Part 4 tests complete.")
    print("=" * 70)

