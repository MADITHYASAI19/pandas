"""Generate sample data for Part 1 (master rollcall) and Part 2 (seating plan).

Creates:
  - data/master_list.xlsx (multi-sheet rollcall workbook with sheets A and B)
  - data/training_seating_sample.xlsx (wide seating plan with merged date headers)
"""

import os
from openpyxl import Workbook

DATA_DIR = "data"
ROSTER_PATH = os.path.join(DATA_DIR, "master_list.xlsx")
SEATING_PATH = os.path.join(DATA_DIR, "training_seating_sample.xlsx")

def build_master_roster():
    wb = Workbook()
    
    # Sheet A
    ws_a = wb.active
    ws_a.title = "A"
    ws_a.append(["Enrollment No", "Name", "Roll No", "Department", "Semester"])
    ws_a.append(["ENR-001", "Aarav Patel", "CS-101", "CS", "Sem 5"])
    ws_a.append(["ENR-002", "Diya Sharma", "CS-102", "CS", "Sem 5"])
    ws_a.append(["ENR-003", "Vihaan Iyer", "IT-201", "IT", "Sem 3"])
    ws_a.append(["ENR-004", "Ananya Rao", "IT-202", "IT", "Sem 3"])
    
    # Sheet B
    ws_b = wb.create_sheet(title="B")
    ws_b.append(["Enrollment No", "Name", "Roll No", "Department", "Semester"])
    ws_b.append(["ENR-005", "Arjun Nair", "ECE-301", "ECE", "Sem 5"])
    ws_b.append(["ENR-006", "Ishita Joshi", "ECE-302", "ECE", "Sem 5"])
    ws_b.append(["ENR-007", "Kabir Verma", "ME-401", "ME", "Sem 3"])
    ws_b.append(["ENR-008", "Saanvi Reddy", "ME-402", "ME", "Sem 3"])
    
    os.makedirs(DATA_DIR, exist_ok=True)
    wb.save(ROSTER_PATH)
    print(f"[OK] Created multi-sheet master roster: {ROSTER_PATH}")


def build_seating_plan():
    wb = Workbook()
    ws = wb.active
    ws.title = "Seating Plan"
    
    # Left fixed headers (Row 1)
    fixed_headers = [
        "Sr No", "Final Batch", "Floor", "Class", "Venue", "Enrollment No",
        "Name", "Gender", "Course", "Branch", "College", "Year of Passing",
        "Parul Mail ID"
    ]
    
    # Let's add the headers for Row 1
    # We will write the fixed headers in columns 1 to 13
    for col_idx, header in enumerate(fixed_headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    # We will add 3 date blocks: 2026-07-01, 2026-07-02, 2026-07-03
    # Each date block is merged across 2 columns
    date_blocks = ["2026-07-01", "2026-07-02", "2026-07-03"]
    for i, date_str in enumerate(date_blocks):
        col_start = 14 + (i * 2)
        # Write date in first column of block
        ws.cell(row=1, column=col_start, value=date_str)
        # Merge first and second columns of this block in row 1
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+1)
        
    # Row 2 headers
    # Fixed columns under row 2 can be empty or repeat
    for col_idx in range(1, 14):
        ws.cell(row=2, column=col_idx, value="")
        
    # Write "Slot 1" and "Slot 2" for each date block in row 2
    for i in range(len(date_blocks)):
        col_start = 14 + (i * 2)
        ws.cell(row=2, column=col_start, value="Slot 1")
        ws.cell(row=2, column=col_start+1, value="Slot 2")
        
    # Student records
    # We have: Sr No, Final Batch, Floor, Class, Venue, Enrollment No, Name, Gender, Course, Branch, College, Year of Passing, Parul Mail ID, followed by Slot 1/2 statuses for the 3 dates (6 statuses total)
    students = [
        ["1", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-001", "Aarav Patel", "Male", "B.Tech", "CS", "PIT", "2027", "aarav@parul.in",
         "Present", "Present", "Present", "Absent", "Present", "Present"],
        ["2", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-002", "Diya Sharma", "Female", "B.Tech", "CS", "PIT", "2027", "diya@parul.in",
         "Present", "Absent", "Present", "Present", "Present", "Leave"],
        ["3", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-003", "Vihaan Iyer", "Male", "B.Tech", "IT", "PIT", "2027", "vihaan@parul.in",
         "Absent", "Absent", "Present", "Present", "Leave", "Present"],
        ["4", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-005", "Arjun Nair", "Male", "B.Tech", "ECE", "PIT", "2027", "arjun@parul.in",
         "Present", "Present", "Leave", "Leave", "Present", "Present"],
        ["5", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-006", "Ishita Joshi", "Female", "B.Tech", "ECE", "PIT", "2027", "ishita@parul.in",
         "Present", "Present", "Present", "Present", "Present", "Present"],
        # ENR-008 is also in roster, we'll write happy values here
        ["6", "Batch 1", "3rd", "Lab 2", "Main Bldg", "ENR-008", "Saanvi Reddy", "Female", "B.Tech", "ME", "PIT", "2027", "saanvi@parul.in",
         "Present", "Present", "Present", "Present", "Present", "Present"],
    ]
    
    for row_idx, student_data in enumerate(students, 3):
        for col_idx, val in enumerate(student_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
            
    wb.save(SEATING_PATH)
    print(f"[OK] Created training seating sample seating plan: {SEATING_PATH}")


if __name__ == "__main__":
    build_master_roster()
    build_seating_plan()
